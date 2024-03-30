from openpyxl import *
import re

def getRowValues(workSheet,row:int):
    values = []
    for column in range(workSheet.max_column):
        values.append(workSheet.cell(row=row,column=column+1).value)
    return values

def getColumnValues(workSheet,column:int):
    values = []
    for row in range(workSheet.max_column):
        values.append(workSheet.cell(row=row+1,column=column))
    return values

def getColumnByName(workSheet, name:str, row:int):
    return getRowValues(workSheet,row).index(name)+1

def getWorkSheetIndexByName(workBook:Workbook,name:str):
    values = []
    for i in workBook.worksheets:
        values.append(i.title)
    return values.index(name)

def listToMultilineString(list:list):
    output = ''
    for i in list:
        output += i + "\n"
    # print(output)
    return output

def multilineStringToList(multiline:str):
    print(re.findall(r"^.+$",multiline,flags=re.MULTILINE))
    return re.findall(r"^.+$",multiline,flags=re.MULTILINE)

def getObjBySearch(query:str,objects:list) -> str:
    searches = re.findall(r'\w+', query)
    print(searches)
    results = []
    titles = getObjectTitles(objects)
    # listAsStr = listToMultilineString(titles)
    for i in range(len(searches)):
        objList = [j for j in objects if searches[i].lower() in j.title.lower()]
        results.extend(objList)
    print(getObjectTitles(results))
    return results

def getObjectTitles(objects:list) -> list:
    output = []
    for i in objects:
        output.append(i.title)
    return output

def getObjects(objectType:list) -> list:
    try:
        output = []
        for i in objectType:
            output.append(i)
        return output
    except ValueError:
        print("getObjects(???) Must be a list of objects")