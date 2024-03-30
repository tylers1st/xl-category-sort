from openpyxl import load_workbook
from openpyxl import *
from utils import *
from window import *
import customtkinter as tk
import io

class Category:
    def __init__(self,id,title,catLvl) -> None:
        self.id = id
        self.title = title
        self.catLvl = catLvl
        pass

class Material:
    def __init__(self,id,partNumber,title,category) -> None:
        self.id = id
        self.id = partNumber
        self.title = title
        self.categories = category
        pass

def updateCatOptions(*args):
    """
    Updates the options menu for categories

    Returns -> None
    """
    #Clears list of buttons in the search
    for widget in catSearchOutputFrame.winfo_children():
        widget.destroy()
    
    searchResultsWidgets = []
        
    # Determines new buttons to add
    if len(usrCatSearch.get()) <= 0:
        selCat.configure(values=getObjectTitles(pbCategories))
        for x in range(len(pbCategories)):
            button = tk.CTkButton(catSearchOutputFrame,width=1,text=pbCategories[x].title,command=print,corner_radius=0)
            searchResultsWidgets.append(button)
            button.grid(sticky="nsew")
        print(searchResultsWidgets[0])
    
    if not len(usrCatSearch.get()) <= 0:
        # selCat['menu'].delete(0,"end")
        searchResults = getObjBySearch(usrCatSearch.get(),pbCategories)
        resultTitles = getObjectTitles(searchResults)
        selCat.configure(values=resultTitles)
        for x in range(len(resultTitles)):
            button = tk.CTkButton(catSearchOutputFrame,width=1,text=resultTitles[x],command=print,corner_radius=0)
            searchResultsWidgets.append(button)
            button.grid(sticky="nsew")
        print(searchResultsWidgets[0])
        
        # optionListStrVar = tk.StringVar(root)
        # optionListStrVar.set(optionList[0])
        # for result in optionList:
        #     varStr = tk.StringVar(root)
        #     varStr.set(value=result)
        #     selCat['menu'].add_command(label=result, command=tk._setit(catPick,result))

def materialSearchOutput(*args):
    if not usrMatSearch.get() == None:
        getObjects()

def clearSearch(*args):
    catSearchBar.delete(0,"end")

# Excel books
pricebook = load_workbook(filename=r'categorysort/src/resources/pricebook.xlsx') # current pricebook
pricebookCatSheet = pricebook['Categories']
pricebookMatSheet = pricebook['Materials']
newPB = Workbook() # new pricebook

topRowPB = getRowValues(pricebookCatSheet,1)
categories = []
categoryColumns = []
pbCategories = []

# Category depth names
for i in range(len(topRowPB)):
    if "Category" in topRowPB[i] and not "Type" in topRowPB[i]:
        categories.append(topRowPB[i])
        categoryColumns.append(i+1)

# Create list of categories as objects
for rowIndex in range(pricebookCatSheet.max_row):
    row = rowIndex+1
    for column in range(len(categories)):
        if not pricebookCatSheet.cell(row=row,column=categoryColumns[column]).value == None:
            pricebookCatSheet.cell(row=row,column=categoryColumns[column]).value
            pbCategories.append(Category(
                pricebookCatSheet.cell(row=row,column=getColumnByName(pricebookCatSheet,"Id",1)).value,
                pricebookCatSheet.cell(row=row,column=categoryColumns[column]).value,
                categories[column]
            ))

# Create list of materials from pricebook as objects
pbMaterials = []
materialColumns = [
    getColumnByName(pricebookMatSheet,"Id",1),
    getColumnByName(pricebookMatSheet,"Code",1),
    getColumnByName(pricebookMatSheet,"Description",1),
    getColumnByName(pricebookMatSheet,"Category.Name",1)
    ]

for rowIndex in range(50):
    row = rowIndex+1
    pbMaterials.append(Material(
        pricebookMatSheet.cell(row=row,column=getColumnByName(pricebookMatSheet,"Id",1)).value,
        pricebookMatSheet.cell(row=row,column=getColumnByName(pricebookMatSheet,"Code",1)).value,
        pricebookMatSheet.cell(row=row,column=getColumnByName(pricebookMatSheet,"Description",1)).value,
        pricebookMatSheet.cell(row=row,column=getColumnByName(pricebookMatSheet,"Category.Name",1)).value
    ))
for i in range(50):
    print(vars(pbMaterials[i]))
    

#Category Search Bar Widgets
usrCatSearch = tk.StringVar()
catSearchBar = tk.CTkEntry(catSearchBarWidgets,textvariable=usrCatSearch)
catSearchBar.grid(row=1,column=1)

cat_search_button = tk.CTkButton(catSearchBarWidgets,text="Search",command=updateCatOptions,width=10,height=10)
cat_search_button.grid(row=1,column=2)
cat_clear_search = tk.CTkButton(catSearchBarWidgets,text="X",command=clearSearch,width=10,height=10)
cat_clear_search.grid(row=1,column=3)

catPick = tk.StringVar()
catPick.set("Select Category")
selCat = tk.CTkOptionMenu(master=catSelectFrame,variable=catPick,values=getObjectTitles(pbCategories))
selCat.pack()
# catOptionsOutput = tk.CTkLabel(master=catSearchOutputFrame,text=listToMultilineString(getObjectTitles(pbCategories)))
# catOptionsOutput.pack()
updateCatOptions()

#Material Select Widgets
usrMatSearch = tk.StringVar()
matSearchBar = tk.CTkEntry(matSearchBarWidgets,textvariable=usrMatSearch)
matSearchBar.grid(row=1,column=1)
mat_search_button = tk.CTkButton(matSearchBarWidgets,text="Search")#TODO add material search feature
mat_search_button.grid(row=1,column=2)

usrMatSearch = tk.StringVar()
matSearchBar = tk.CTkEntry(matSearchBarWidgets,textvariable=usrMatSearch)
matSearchBar.grid(column=2,row=1)

mat_search_button = tk.CTkButton(matSearchBarWidgets,text="Search",command=materialSearchOutput)
mat_search_button.grid(column=2,row=1)



# fileOutput = io.open(r"categorysort\src\resources\output.txt",'w')
# fileOutput.write(listToMultilineString(getObjectTitles(pbCategories)))
# fileOutput.close

root.mainloop()